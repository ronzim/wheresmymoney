from __future__ import annotations

from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook

from wheresmymoney.models import Transaction


class ParserError(ValueError):
    pass


STRUCTURED_XLSX_HEADERS = (
    "Data Op.",
    "Data Val.",
    "Causale",
    "Descrizione",
    "Importo",
    "Div.",
)

SPLIT_AMOUNT_XLSX_HEADERS = (
    "Data Contabile",
    "Data Valuta",
    "Addebiti (euro)",
    "Accrediti (euro)",
    "Descrizione operazioni",
)

HTML_XLS_HEADERS = (
    "Data Contabile",
    "Data Valuta",
    "Importo",
    "Divisa",
    "Causale / Descrizione",
    "Canale",
)


@dataclass(frozen=True)
class ParsedStatement:
    source_bank: str
    transactions: tuple[Transaction, ...]
    parser_name: str


def parse_statement(file_path: str | Path, source_bank: str) -> ParsedStatement:
    path = Path(file_path)
    parser_name = detect_parser(path)

    if parser_name == "structured_xlsx":
        transactions = tuple(parse_structured_xlsx(path, source_bank))
    elif parser_name == "split_amount_xlsx":
        transactions = tuple(parse_split_amount_xlsx(path, source_bank))
    elif parser_name == "html_xls":
        transactions = tuple(parse_html_xls(path, source_bank))
    else:  # pragma: no cover - guarded by detect_parser
        raise ParserError(f"Unsupported parser name: {parser_name}")

    if not transactions:
        raise ParserError(f"No transactions found in file: {path}")

    return ParsedStatement(
        source_bank=source_bank,
        transactions=transactions,
        parser_name=parser_name,
    )


def detect_parser(file_path: str | Path) -> str:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        if _has_structured_xlsx_header(path):
            return "structured_xlsx"
        if _has_split_amount_xlsx_header(path):
            return "split_amount_xlsx"
        raise ParserError(f"Unsupported XLSX layout: {path.name}")

    if suffix == ".xls":
        if _looks_like_html_statement(path):
            return "html_xls"
        raise ParserError(f"Unsupported XLS layout: {path.name}")

    raise ParserError(f"Unsupported file extension: {path.suffix}")


def parse_structured_xlsx(file_path: str | Path, source_bank: str) -> list[Transaction]:
    path = Path(file_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)

    header_row_index = None
    for row_index, row in enumerate(rows, start=1):
        compact = tuple(row[:6])
        if compact == STRUCTURED_XLSX_HEADERS:
            header_row_index = row_index
            break

    if header_row_index is None:
        raise ParserError(f"Structured XLSX header not found in {path.name}")

    transactions: list[Transaction] = []
    for row in rows:
        data_op, data_val, _causale, descrizione, importo, divisa = row[:6]
        if all(value in (None, "") for value in (data_op, data_val, descrizione, importo, divisa)):
            continue
        transactions.append(
            Transaction(
                source_bank=source_bank,
                transaction_date=str(data_op),
                value_date=str(data_val),
                amount=str(importo),
                currency=str(divisa),
                original_description=str(descrizione),
            )
        )

    return transactions


def parse_split_amount_xlsx(file_path: str | Path, source_bank: str) -> list[Transaction]:
    path = Path(file_path)
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row_index = None
    for row_index in range(1, worksheet.max_row + 1):
        row = tuple(
            worksheet.cell(row_index, column_index).value
            for column_index in range(1, 6)
        )
        if row == SPLIT_AMOUNT_XLSX_HEADERS:
            header_row_index = row_index
            break

    if header_row_index is None:
        raise ParserError(f"Split-amount XLSX header not found in {path.name}")

    transactions: list[Transaction] = []
    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        data_contabile, data_valuta, addebiti, accrediti, descrizione = (
            worksheet.cell(row_index, column_index).value
            for column_index in range(1, 6)
        )
        if all(
            value in (None, "")
            for value in (data_contabile, data_valuta, addebiti, accrediti, descrizione)
        ):
            continue

        amount = _signed_amount_from_split_columns(addebiti, accrediti, path.name)
        transactions.append(
            Transaction(
                source_bank=source_bank,
                transaction_date=data_contabile,
                value_date=data_valuta,
                amount=amount,
                currency="EUR",
                original_description=str(descrizione),
            )
        )

    return transactions


def parse_html_xls(file_path: str | Path, source_bank: str) -> list[Transaction]:
    path = Path(file_path)
    parser = _StatementTableParser()
    parser.feed(path.read_text(encoding="utf-8", errors="ignore"))
    rows = parser.rows

    if not rows:
        raise ParserError(f"No HTML table rows found in {path.name}")

    header = tuple(rows[0])
    if header != HTML_XLS_HEADERS:
        raise ParserError(f"Unexpected HTML XLS header in {path.name}: {header}")

    transactions: list[Transaction] = []
    for row in rows[1:]:
        if len(row) < 5:
            continue
        data_contabile, data_valuta, importo, divisa, descrizione = row[:5]
        if not any(cell.strip() for cell in row if isinstance(cell, str)):
            continue
        transactions.append(
            Transaction(
                source_bank=source_bank,
                transaction_date=data_contabile,
                value_date=data_valuta,
                amount=importo,
                currency=divisa,
                original_description=descrizione,
            )
        )

    return transactions


def _has_structured_xlsx_header(path: Path) -> bool:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    for row in worksheet.iter_rows(values_only=True):
        if tuple(row[:6]) == STRUCTURED_XLSX_HEADERS:
            return True
    return False


def _has_split_amount_xlsx_header(path: Path) -> bool:
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    for row_index in range(1, worksheet.max_row + 1):
        row = tuple(worksheet.cell(row_index, column_index).value for column_index in range(1, 6))
        if row == SPLIT_AMOUNT_XLSX_HEADERS:
            return True
    return False


def _looks_like_html_statement(path: Path) -> bool:
    content = path.read_text(encoding="utf-8", errors="ignore").lower()
    return (
        "<html" in content
        and "causale / descrizione" in content
        and 'table id="ccmo"' in content
    )


def _signed_amount_from_split_columns(
    debit: object,
    credit: object,
    file_name: str,
) -> str:
    debit_present = debit not in (None, "")
    credit_present = credit not in (None, "")

    if debit_present and credit_present:
        raise ParserError(
            f"Row in {file_name} contains both debit and credit values"
        )
    if not debit_present and not credit_present:
        raise ParserError(
            f"Row in {file_name} contains neither debit nor credit value"
        )

    if debit_present:
        return f"-{debit}"
    return str(credit)


class _StatementTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None
        self._in_relevant_table = False
        self._table_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = dict(attrs)
        if tag == "table" and attrs_dict.get("id") == "CCMO":
            self._in_relevant_table = True
            self._table_depth = 1
            return

        if not self._in_relevant_table:
            return

        if tag == "table":
            self._table_depth += 1
        elif tag == "tr":
            self._current_row = []
        elif tag in {"td", "th"}:
            self._current_cell = []

    def handle_endtag(self, tag: str) -> None:
        if not self._in_relevant_table:
            return

        if tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            value = "".join(self._current_cell)
            value = value.replace("\xa0", " ")
            self._current_row.append(value)
            self._current_cell = None
            return

        if tag == "tr" and self._current_row is not None:
            if any(cell.strip() for cell in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None
            return

        if tag == "table":
            self._table_depth -= 1
            if self._table_depth == 0:
                self._in_relevant_table = False

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)